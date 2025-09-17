from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import os
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as OpenPyXLImage
import io

app = Flask(__name__)
app.config['SECRET_KEY'] = 'tu_clave_secreta_aqui_cambiala_en_produccion'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///taller.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Modelos de la base de datos
class Usuario(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    nombre_completo = db.Column(db.String(120), nullable=False)
    rol = db.Column(db.String(20), default='usuario')
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)

class Vehiculo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    marca = db.Column(db.String(50), nullable=False)
    modelo = db.Column(db.String(50), nullable=False)
    año = db.Column(db.Integer, nullable=False)
    color = db.Column(db.String(30), nullable=False)
    placas = db.Column(db.String(20), unique=True, nullable=False)
    vin = db.Column(db.String(17), unique=True)
    cliente_nombre = db.Column(db.String(100), nullable=False)
    cliente_telefono = db.Column(db.String(20))
    cliente_email = db.Column(db.String(100))
    fecha_entrada = db.Column(db.DateTime, default=datetime.utcnow)
    fecha_salida = db.Column(db.DateTime)
    estado = db.Column(db.String(20), default='en_taller')  # en_taller, terminado, entregado
    observaciones = db.Column(db.Text)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    
    # Relaciones
    usuario = db.relationship('Usuario', backref='vehiculos')

class Refaccion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), unique=True, nullable=False)
    nombre = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.Text)
    precio = db.Column(db.Float, nullable=False)
    stock = db.Column(db.Integer, default=0)
    categoria = db.Column(db.String(50))
    proveedor = db.Column(db.String(100))

class Cotizacion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    vehiculo_id = db.Column(db.Integer, db.ForeignKey('vehiculo.id'), nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    subtotal = db.Column(db.Float, default=0.0)
    iva = db.Column(db.Float, default=0.0)
    total = db.Column(db.Float, default=0.0)
    estado = db.Column(db.String(20), default='pendiente')  # pendiente, aprobada, rechazada
    observaciones = db.Column(db.Text)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    
    # Relaciones
    vehiculo = db.relationship('Vehiculo', backref='cotizaciones')
    usuario = db.relationship('Usuario', backref='cotizaciones')
    cotizacion_refacciones = db.relationship('CotizacionRefaccion', backref='cotizacion', cascade='all, delete-orphan')

class CotizacionRefaccion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cotizacion_id = db.Column(db.Integer, db.ForeignKey('cotizacion.id'), nullable=False)
    refaccion_id = db.Column(db.Integer, db.ForeignKey('refaccion.id'), nullable=False)
    cantidad = db.Column(db.Integer, nullable=False)
    precio_unitario = db.Column(db.Float, nullable=False)
    subtotal = db.Column(db.Float, nullable=False)
    
    # Relaciones
    refaccion = db.relationship('Refaccion', backref='cotizacion_refacciones')

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

# Rutas principales
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = Usuario.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            flash('¡Bienvenido!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Has cerrado sesión', 'info')
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    vehiculos_en_taller = Vehiculo.query.filter_by(estado='en_taller').count()
    vehiculos_terminados = Vehiculo.query.filter_by(estado='terminado').count()
    total_vehiculos = Vehiculo.query.count()
    
    # Vehículos recientes
    vehiculos_recientes = Vehiculo.query.order_by(Vehiculo.fecha_entrada.desc()).limit(5).all()
    
    return render_template('dashboard.html', 
                         vehiculos_en_taller=vehiculos_en_taller,
                         vehiculos_terminados=vehiculos_terminados,
                         total_vehiculos=total_vehiculos,
                         vehiculos_recientes=vehiculos_recientes)

# Rutas para vehículos
@app.route('/vehiculos')
@login_required
def vehiculos():
    vehiculos = Vehiculo.query.order_by(Vehiculo.fecha_entrada.desc()).all()
    return render_template('vehiculos.html', vehiculos=vehiculos)

@app.route('/vehiculos/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_vehiculo():
    if request.method == 'POST':
        vehiculo = Vehiculo(
            marca=request.form['marca'],
            modelo=request.form['modelo'],
            año=int(request.form['año']),
            color=request.form['color'],
            placas=request.form['placas'],
            vin=request.form['vin'],
            cliente_nombre=request.form['cliente_nombre'],
            cliente_telefono=request.form['cliente_telefono'],
            cliente_email=request.form['cliente_email'],
            observaciones=request.form['observaciones'],
            usuario_id=current_user.id
        )
        db.session.add(vehiculo)
        db.session.commit()
        flash('Vehículo registrado exitosamente', 'success')
        return redirect(url_for('vehiculos'))
    
    return render_template('nuevo_vehiculo.html')

@app.route('/vehiculos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_vehiculo(id):
    vehiculo = Vehiculo.query.get_or_404(id)
    if request.method == 'POST':
        vehiculo.marca = request.form['marca']
        vehiculo.modelo = request.form['modelo']
        vehiculo.año = int(request.form['año'])
        vehiculo.color = request.form['color']
        vehiculo.placas = request.form['placas']
        vehiculo.vin = request.form['vin']
        vehiculo.cliente_nombre = request.form['cliente_nombre']
        vehiculo.cliente_telefono = request.form['cliente_telefono']
        vehiculo.cliente_email = request.form['cliente_email']
        vehiculo.observaciones = request.form['observaciones']
        vehiculo.estado = request.form['estado']
        
        # Procesar fecha de entrada
        if request.form.get('fecha_entrada'):
            try:
                fecha_entrada_str = request.form['fecha_entrada']
                vehiculo.fecha_entrada = datetime.strptime(fecha_entrada_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                flash('Formato de fecha de entrada inválido', 'error')
                return render_template('editar_vehiculo.html', vehiculo=vehiculo)
        
        # Procesar fecha de salida
        if request.form.get('fecha_salida'):
            try:
                fecha_salida_str = request.form['fecha_salida']
                vehiculo.fecha_salida = datetime.strptime(fecha_salida_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                flash('Formato de fecha de salida inválido', 'error')
                return render_template('editar_vehiculo.html', vehiculo=vehiculo)
        else:
            # Si no hay fecha de salida y el estado es entregado, establecer fecha actual
            if vehiculo.estado == 'entregado' and not vehiculo.fecha_salida:
                vehiculo.fecha_salida = datetime.utcnow()
            elif vehiculo.estado != 'entregado':
                vehiculo.fecha_salida = None
        
        db.session.commit()
        flash('Vehículo actualizado exitosamente', 'success')
        return redirect(url_for('vehiculos'))
    
    return render_template('editar_vehiculo.html', vehiculo=vehiculo)

# Rutas para refacciones
@app.route('/refacciones')
@login_required
def refacciones():
    refacciones = Refaccion.query.order_by(Refaccion.nombre).all()
    return render_template('refacciones.html', refacciones=refacciones)

@app.route('/refacciones/nueva', methods=['GET', 'POST'])
@login_required
def nueva_refaccion():
    if request.method == 'POST':
        refaccion = Refaccion(
            codigo=request.form['codigo'],
            nombre=request.form['nombre'],
            descripcion=request.form['descripcion'],
            precio=float(request.form['precio']),
            stock=int(request.form['stock']),
            categoria=request.form['categoria'],
            proveedor=request.form['proveedor']
        )
        db.session.add(refaccion)
        db.session.commit()
        flash('Refacción agregada exitosamente', 'success')
        return redirect(url_for('refacciones'))
    
    return render_template('nueva_refaccion.html')

# Rutas para cotizaciones
@app.route('/cotizaciones')
@login_required
def cotizaciones():
    cotizaciones = Cotizacion.query.order_by(Cotizacion.fecha_creacion.desc()).all()
    return render_template('cotizaciones.html', cotizaciones=cotizaciones)

@app.route('/cotizaciones/nueva/<int:vehiculo_id>', methods=['GET', 'POST'])
@login_required
def nueva_cotizacion(vehiculo_id):
    vehiculo = Vehiculo.query.get_or_404(vehiculo_id)
    refacciones = Refaccion.query.all()
    
    if request.method == 'POST':
        cotizacion = Cotizacion(
            vehiculo_id=vehiculo_id,
            observaciones=request.form['observaciones'],
            usuario_id=current_user.id
        )
        db.session.add(cotizacion)
        db.session.flush()  # Para obtener el ID de la cotización
        
        # Procesar refacciones seleccionadas
        refacciones_ids = request.form.getlist('refacciones[]')
        cantidades = request.form.getlist('cantidades[]')
        
        subtotal = 0
        for i, refaccion_id in enumerate(refacciones_ids):
            if refaccion_id and cantidades[i]:
                refaccion = Refaccion.query.get(refaccion_id)
                cantidad = int(cantidades[i])
                precio_unitario = refaccion.precio
                subtotal_item = cantidad * precio_unitario
                
                cotizacion_refaccion = CotizacionRefaccion(
                    cotizacion_id=cotizacion.id,
                    refaccion_id=refaccion_id,
                    cantidad=cantidad,
                    precio_unitario=precio_unitario,
                    subtotal=subtotal_item
                )
                db.session.add(cotizacion_refaccion)
                subtotal += subtotal_item
        
        cotizacion.subtotal = subtotal
        cotizacion.iva = subtotal * 0.16  # 16% IVA
        cotizacion.total = subtotal + cotizacion.iva
        
        db.session.commit()
        flash('Cotización creada exitosamente', 'success')
        return redirect(url_for('cotizaciones'))
    
    return render_template('nueva_cotizacion.html', vehiculo=vehiculo, refacciones=refacciones)

@app.route('/cotizaciones/<int:id>/pdf')
@login_required
def cotizacion_pdf(id):
    cotizacion = Cotizacion.query.get_or_404(id)
    
    # Crear PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centrado
    )
    
    # Logo y encabezado
    try:
        logo_path = os.path.join(app.static_folder, 'images', 'myLogo.png')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=1.5*inch, height=1*inch)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 10))
    except:
        pass  # Si no se puede cargar el logo, continuar sin él
    
    # Título
    elements.append(Paragraph("MANTENIMIENTO BMW/MINI TOLUCA Y METEPEC", title_style))
    elements.append(Paragraph("COTIZACIÓN DE SERVICIOS", title_style))
    elements.append(Spacer(1, 20))
    
    # Información del vehículo
    vehiculo = cotizacion.vehiculo
    elements.append(Paragraph(f"<b>Vehículo:</b> {vehiculo.marca} {vehiculo.modelo} {vehiculo.año}", styles['Normal']))
    elements.append(Paragraph(f"<b>Placas:</b> {vehiculo.placas}", styles['Normal']))
    elements.append(Paragraph(f"<b>Cliente:</b> {vehiculo.cliente_nombre}", styles['Normal']))
    elements.append(Paragraph(f"<b>Fecha:</b> {cotizacion.fecha_creacion.strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Tabla de refacciones
    data = [['Refacción', 'Cantidad', 'Precio Unit.', 'Subtotal']]
    for item in cotizacion.cotizacion_refacciones:
        refaccion = item.refaccion
        data.append([
            refaccion.nombre,
            str(item.cantidad),
            f"${item.precio_unitario:.2f}",
            f"${item.subtotal:.2f}"
        ])
    
    # Totales
    data.append(['', '', 'Subtotal:', f"${cotizacion.subtotal:.2f}"])
    data.append(['', '', 'IVA (16%):', f"${cotizacion.iva:.2f}"])
    data.append(['', '', 'Total:', f"${cotizacion.total:.2f}"])
    
    table = Table(data, colWidths=[3*inch, 1*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -4), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (2, -3), (-1, -1), 'RIGHT'),
        ('FONTNAME', (2, -3), (-1, -1), 'Helvetica-Bold'),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    if cotizacion.observaciones:
        elements.append(Paragraph(f"<b>Observaciones:</b> {cotizacion.observaciones}", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'cotizacion_{cotizacion.id}.pdf',
        mimetype='application/pdf'
    )

@app.route('/cotizaciones/<int:id>/excel')
@login_required
def cotizacion_excel(id):
    cotizacion = Cotizacion.query.get_or_404(id)
    vehiculo = cotizacion.vehiculo
    
    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotización"
    
    # Estilos
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Logo
    try:
        logo_path = os.path.join(app.static_folder, 'images', 'myLogo.png')
        if os.path.exists(logo_path):
            img = OpenPyXLImage(logo_path)
            img.width = 120
            img.height = 80
            ws.add_image(img, 'A1')
    except:
        pass  # Si no se puede cargar el logo, continuar sin él
    
    # Título
    ws['A4'] = "MANTENIMIENTO BMW/MINI TOLUCA Y METEPEC"
    ws['A4'].font = title_font
    ws.merge_cells('A4:D4')
    ws['A4'].alignment = Alignment(horizontal='center')
    
    ws['A5'] = "COTIZACIÓN DE SERVICIOS"
    ws['A5'].font = title_font
    ws.merge_cells('A5:D5')
    ws['A5'].alignment = Alignment(horizontal='center')
    
    # Información del vehículo
    ws['A7'] = f"Vehículo: {vehiculo.marca} {vehiculo.modelo} {vehiculo.año}"
    ws['A8'] = f"Placas: {vehiculo.placas}"
    ws['A9'] = f"Cliente: {vehiculo.cliente_nombre}"
    ws['A10'] = f"Fecha: {cotizacion.fecha_creacion.strftime('%d/%m/%Y')}"
    
    # Encabezados de tabla
    headers = ['Refacción', 'Cantidad', 'Precio Unit.', 'Subtotal']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=13, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Datos de refacciones
    row = 14
    for item in cotizacion.cotizacion_refacciones:
        refaccion = item.refaccion
        ws.cell(row=row, column=1, value=refaccion.nombre)
        ws.cell(row=row, column=2, value=item.cantidad)
        ws.cell(row=row, column=3, value=item.precio_unitario)
        ws.cell(row=row, column=4, value=item.subtotal)
        row += 1
    
    # Totales
    total_row = row + 1
    ws.cell(row=total_row, column=3, value="Subtotal:").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=cotizacion.subtotal)
    
    ws.cell(row=total_row + 1, column=3, value="IVA (16%):").font = Font(bold=True)
    ws.cell(row=total_row + 1, column=4, value=cotizacion.iva)
    
    ws.cell(row=total_row + 2, column=3, value="Total:").font = Font(bold=True)
    ws.cell(row=total_row + 2, column=4, value=cotizacion.total)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'cotizacion_{cotizacion.id}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# API para obtener refacciones
@app.route('/api/refacciones')
@login_required
def api_refacciones():
    refacciones = Refaccion.query.all()
    return jsonify([{
        'id': r.id,
        'codigo': r.codigo,
        'nombre': r.nombre,
        'precio': r.precio,
        'stock': r.stock
    } for r in refacciones])

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # Crear usuario administrador si no existe
        admin = Usuario.query.filter_by(username='admin').first()
        if not admin:
            admin = Usuario(
                username='admin',
                email='admin@taller.com',
                password_hash=generate_password_hash('admin123'),
                nombre_completo='Administrador',
                rol='admin'
            )
            db.session.add(admin)
            db.session.commit()
            print("Usuario administrador creado: admin / admin123")
    
    app.run(debug=True, host='0.0.0.0', port=5000) 